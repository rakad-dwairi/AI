from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class RubricCriterion:
    criterion_id: str
    axis: str
    name: str
    evaluation_method: str | None
    metric: str | None
    equation: str | None
    reference_scale: str | None
    source_row: int


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _slug(text: str, fallback: str) -> str:
    ascii_hint = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    if ascii_hint:
        return ascii_hint[:80]
    return fallback


def load_form_criteria(form_path: Path) -> list[RubricCriterion]:
    """Extract machine-readable criteria from the tool-capture sheet."""
    workbook = load_workbook(form_path, data_only=True)
    worksheet = workbook["رصد الأدء الظهوري بالTool"]
    criteria: list[RubricCriterion] = []
    current_axis = ""
    current_name = ""

    for row_index in range(11, worksheet.max_row + 1):
        axis = _clean(worksheet.cell(row_index, 4).value)
        name = _clean(worksheet.cell(row_index, 5).value)
        method = _clean(worksheet.cell(row_index, 6).value)
        metric = _clean(worksheet.cell(row_index, 7).value)
        equation = _clean(worksheet.cell(row_index, 9).value)
        reference = _clean(worksheet.cell(row_index, 10).value)

        if axis:
            current_axis = axis
        if name:
            current_name = name

        if not any([method, metric, equation, reference]) or not current_name:
            continue

        row_key = _slug(f"{current_axis}_{current_name}_{row_index}", f"criterion_{row_index}")
        criteria.append(
            RubricCriterion(
                criterion_id=row_key,
                axis=current_axis,
                name=current_name,
                evaluation_method=method,
                metric=metric,
                equation=equation,
                reference_scale=reference,
                source_row=row_index,
            )
        )

    return criteria


def load_rubric_levels(form_path: Path) -> list[dict[str, str | int | None]]:
    workbook = load_workbook(form_path, data_only=True)
    worksheet = workbook["روبريك رصد الأداء الظهوري"]
    rows: list[dict[str, str | int | None]] = []
    current_axis = ""

    for row_index in range(5, worksheet.max_row + 1):
        axis = _clean(worksheet.cell(row_index, 4).value) or _clean(worksheet.cell(row_index, 5).value)
        item = _clean(worksheet.cell(row_index, 6).value)
        if axis:
            current_axis = axis
        if not item:
            continue
        rows.append(
            {
                "source_row": row_index,
                "axis": current_axis,
                "item": item,
                "level_5": _clean(worksheet.cell(row_index, 7).value),
                "level_4": _clean(worksheet.cell(row_index, 8).value),
                "level_3": _clean(worksheet.cell(row_index, 9).value),
                "level_2": _clean(worksheet.cell(row_index, 10).value),
                "level_1": _clean(worksheet.cell(row_index, 11).value),
            }
        )
    return rows
